"""
Excel report generator using openpyxl.
"""
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    NamedStyle, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import (
    BarChart, Reference, Series, PieChart
)
from openpyxl.drawing.image import Image
from typing import Dict, List, Any, Optional, Tuple
import os
from datetime import datetime
from pathlib import Path
import json
import tempfile

from shared.utils.logging import logger
from services.reporting_service.config import config


class ExcelGenerator:
    """Generates Excel audit reports with multiple sheets."""
    
    def __init__(self):
        self.workbook = None
        self.styles = {}
        self._create_styles()
        
    def _create_styles(self):
        """Create named styles for Excel."""
        # This will be populated when workbook is created
        pass
    
    def _init_styles(self, workbook):
        """Initialize named styles in workbook."""
        # Title style
        title_style = NamedStyle(name="title_style")
        title_style.font = Font(name='Calibri', size=16, bold=True, color='2C3E50')
        title_style.alignment = Alignment(horizontal='center', vertical='center')
        workbook.add_named_style(title_style)
        
        # Header style
        header_style = NamedStyle(name="header_style")
        header_style.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        header_style.fill = PatternFill(start_color='3498DB', end_color='3498DB', fill_type='solid')
        header_style.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        workbook.add_named_style(header_style)
        
        # Critical finding style
        critical_style = NamedStyle(name="critical_style")
        critical_style.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        critical_style.fill = PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid')
        critical_style.alignment = Alignment(horizontal='center', vertical='center')
        workbook.add_named_style(critical_style)
        
        # High finding style
        high_style = NamedStyle(name="high_style")
        high_style.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        high_style.fill = PatternFill(start_color='E67E22', end_color='E67E22', fill_type='solid')
        high_style.alignment = Alignment(horizontal='center', vertical='center')
        workbook.add_named_style(high_style)
        
        # Medium finding style
        medium_style = NamedStyle(name="medium_style")
        medium_style.font = Font(name='Calibri', size=10, bold=True, color='000000')
        medium_style.fill = PatternFill(start_color='F1C40F', end_color='F1C40F', fill_type='solid')
        medium_style.alignment = Alignment(horizontal='center', vertical='center')
        workbook.add_named_style(medium_style)
        
        # Low finding style
        low_style = NamedStyle(name="low_style")
        low_style.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        low_style.fill = PatternFill(start_color='2ECC71', end_color='2ECC71', fill_type='solid')
        low_style.alignment = Alignment(horizontal='center', vertical='center')
        workbook.add_named_style(low_style)
        
        # Normal cell style
        normal_style = NamedStyle(name="normal_style")
        normal_style.font = Font(name='Calibri', size=10)
        normal_style.alignment = Alignment(vertical='top', wrap_text=True)
        workbook.add_named_style(normal_style)
        
        # Code cell style
        code_style = NamedStyle(name="code_style")
        code_style.font = Font(name='Courier New', size=9)
        code_style.fill = PatternFill(start_color='F8F9F9', end_color='F8F9F9', fill_type='solid')
        code_style.alignment = Alignment(vertical='top', wrap_text=True)
        workbook.add_named_style(code_style)
        
        # Date style
        date_style = NamedStyle(name="date_style")
        date_style.font = Font(name='Calibri', size=10)
        date_style.number_format = 'YYYY-MM-DD HH:MM'
        workbook.add_named_style(date_style)
        
        # Footer style
        footer_style = NamedStyle(name="footer_style")
        footer_style.font = Font(name='Calibri', size=9, italic=True, color='7F8C8D')
        workbook.add_named_style(footer_style)
        
        self.styles = {
            'title': title_style,
            'header': header_style,
            'critical': critical_style,
            'high': high_style,
            'medium': medium_style,
            'low': low_style,
            'normal': normal_style,
            'code': code_style,
            'date': date_style,
            'footer': footer_style
        }
    
    def generate_report(
        self,
        report_data: Dict[str, Any],
        output_path: str,
        include_charts: bool = True
    ) -> Dict[str, Any]:
        """
        Generate Excel report.
        
        Args:
            report_data: Report data
            output_path: Output file path
            include_charts: Whether to include charts
        
        Returns:
            Generation metadata
        """
        start_time = datetime.now()
        
        try:
            # Create workbook
            self.workbook = Workbook()
            self._init_styles(self.workbook)
            
            # Remove default sheet
            default_sheet = self.workbook.active
            self.workbook.remove(default_sheet)
            
            # Create sheets
            self._create_cover_sheet(report_data)
            self._create_executive_summary_sheet(report_data)
            self._create_findings_sheet(report_data)
            
            if 'recommendations' in report_data:
                self._create_recommendations_sheet(report_data)
            
            if 'appendix' in report_data:
                self._create_appendix_sheet(report_data)
            
            # Add summary dashboard if charts are enabled
            if include_charts and config.excel_include_charts:
                self._create_dashboard_sheet(report_data)
            
            # Save workbook
            self.workbook.save(output_path)
            
            # Calculate metadata
            end_time = datetime.now()
            duration = (end_time - start_time).total_seconds()
            
            # Get file size
            file_size = os.path.getsize(output_path)
            
            metadata = {
                'success': True,
                'output_path': output_path,
                'file_size': file_size,
                'generation_time': duration,
                'sheet_count': len(self.workbook.sheetnames),
                'format': 'excel',
                'timestamp': datetime.now().isoformat()
            }
            
            logger.info(f"Excel report generated: {output_path} ({file_size} bytes)")
            
            return metadata
            
        except Exception as e:
            logger.error(f"Error generating Excel report: {str(e)}")
            raise
        finally:
            self.workbook = None
    
    def _create_cover_sheet(self, report_data: Dict[str, Any]):
        """Create cover sheet."""
        sheet = self.workbook.create_sheet(title="Cover")
        
        # Set column widths
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 40
        
        # Title
        sheet['A1'] = config.company_name
        sheet['A1'].style = self.styles['title']
        sheet.merge_cells('A1:B1')
        
        sheet['A3'] = "AUDIT REPORT"
        sheet['A3'].style = self.styles['title']
        sheet.merge_cells('A3:B3')
        
        # Report details
        row = 5
        
        details = [
            ("Report ID:", report_data.get('report_id', 'N/A')),
            ("Report Title:", report_data.get('title', 'Audit Report')),
            ("Generated:", report_data.get('generated_date', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))),
            ("Period:", report_data.get('period', 'N/A')),
            ("Scope:", report_data.get('scope', 'N/A')),
            ("Generated By:", report_data.get('generated_by', 'Aquila Audit System')),
        ]
        
        # Add tenant info if available
        if 'tenant' in report_data:
            details.append(("Tenant:", report_data['tenant'].get('name', 'N/A')))
        
        for label, value in details:
            sheet[f'A{row}'] = label
            sheet[f'A{row}'].font = Font(bold=True)
            sheet[f'B{row}'] = value
            row += 1
        
        # Confidential notice
        if report_data.get('confidential', True):
            row += 1
            sheet[f'A{row}'] = "CONFIDENTIAL"
            sheet[f'A{row}'].font = Font(bold=True, color='E74C3C')
            sheet.merge_cells(f'A{row}:B{row}')
            
            row += 1
            sheet[f'A{row}'] = "This report contains sensitive information. Distribution is restricted to authorized personnel only."
            sheet.merge_cells(f'A{row}:B{row}')
            sheet[f'A{row}'].alignment = Alignment(wrap_text=True)
        
        # Footer
        max_row = sheet.max_row + 3
        sheet[f'A{max_row}'] = config.company_name
        sheet[f'A{max_row}'].style = self.styles['footer']
        
        sheet[f'A{max_row + 1}'] = config.company_address
        sheet[f'A{max_row + 1}'].style = self.styles['footer']
        
        sheet[f'A{max_row + 2}'] = config.company_website
        sheet[f'A{max_row + 2}'].style = self.styles['footer']
        
        # Add borders to main area
        for row in range(5, row):
            for col in range(1, 3):
                cell = sheet.cell(row=row, column=col)
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.border = thin_border
    
    def _create_executive_summary_sheet(self, report_data: Dict[str, Any]):
        """Create executive summary sheet."""
        sheet = self.workbook.create_sheet(title="Executive Summary")
        
        # Title
        sheet['A1'] = "Executive Summary"
        sheet['A1'].style = self.styles['title']
        sheet.merge_cells('A1:F1')
        
        row = 3
        
        # Summary text
        if 'executive_summary' in report_data:
            sheet[f'A{row}'] = "Summary:"
            sheet[f'A{row}'].font = Font(bold=True)
            
            summary = report_data['executive_summary']
            sheet[f'B{row}'] = summary
            sheet[f'B{row}'].style = self.styles['normal']
            sheet.merge_cells(f'B{row}:F{row}')
            sheet.row_dimensions[row].height = 60
            row += 2
        
        # Key metrics
        metrics = report_data.get('metrics', {})
        if metrics:
            sheet[f'A{row}'] = "Key Metrics"
            sheet[f'A{row}'].font = Font(size=12, bold=True, color='2C3E50')
            sheet.merge_cells(f'A{row}:F{row}')
            row += 1
            
            # Metrics table headers
            headers = ["Metric", "Value", "Description"]
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=row, column=col)
                cell.value = header
                cell.style = self.styles['header']
            
            row += 1
            
            # Define metrics to display
            metric_definitions = [
                ("Total Findings", metrics.get('total_findings', 0), "Total number of findings identified"),
                ("Critical Findings", metrics.get('critical', 0), "Findings requiring immediate attention"),
                ("High Findings", metrics.get('high', 0), "High priority findings"),
                ("Medium Findings", metrics.get('medium', 0), "Medium priority findings"),
                ("Low Findings", metrics.get('low', 0), "Low priority findings"),
            ]
            
            if 'risk_score' in metrics:
                metric_definitions.append(
                    ("Risk Score", f"{metrics['risk_score']}/10", "Overall risk assessment score")
                )
            
            for metric_name, value, description in metric_definitions:
                sheet.cell(row=row, column=1, value=metric_name)
                sheet.cell(row=row, column=2, value=value)
                sheet.cell(row=row, column=3, value=description)
                
                # Style the row
                for col in range(1, 4):
                    cell = sheet.cell(row=row, column=col)
                    cell.style = self.styles['normal']
                    cell.border = Border(
                        bottom=Side(style='thin', color='D5D8DC')
                    )
                
                row += 1
            
            row += 1
        
        # Top findings
        top_findings = report_data.get('top_findings', [])[:5]
        if top_findings:
            sheet[f'A{row}'] = "Top Findings"
            sheet[f'A{row}'].font = Font(size=12, bold=True, color='2C3E50')
            sheet.merge_cells(f'A{row}:F{row}')
            row += 1
            
            # Findings table headers
            headers = ["#", "Title", "Severity", "Rule", "Description"]
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=row, column=col)
                cell.value = header
                cell.style = self.styles['header']
            
            row += 1
            
            for i, finding in enumerate(top_findings, 1):
                severity = finding.get('severity', 'Medium').upper()
                
                # Set cell values
                sheet.cell(row=row, column=1, value=i)
                sheet.cell(row=row, column=2, value=finding.get('title', 'Finding'))
                sheet.cell(row=row, column=3, value=severity)
                sheet.cell(row=row, column=4, value=finding.get('rule_name', 'N/A'))
                sheet.cell(row=row, column=5, value=finding.get('description', ''))
                
                # Apply severity-specific style
                severity_cell = sheet.cell(row=row, column=3)
                severity_style = self._get_severity_style(severity)
                severity_cell.style = severity_style
                
                # Style other cells
                for col in [1, 2, 4, 5]:
                    cell = sheet.cell(row=row, column=col)
                    cell.style = self.styles['normal']
                    cell.border = Border(
                        bottom=Side(style='thin', color='D5D8DC')
                    )
                    if col == 5:  # Description column
                        sheet.column_dimensions[get_column_letter(col)].width = 60
                
                row += 1
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_findings_sheet(self, report_data: Dict[str, Any]):
        """Create detailed findings sheet."""
        sheet = self.workbook.create_sheet(title="Findings")
        
        # Title
        sheet['A1'] = "Detailed Findings"
        sheet['A1'].style = self.styles['title']
        sheet.merge_cells('A1:J1')
        
        row = 3
        
        findings = report_data.get('findings', [])
        
        if not findings:
            sheet['A3'] = "No findings to report."
            sheet['A3'].style = self.styles['normal']
            return
        
        # Table headers
        headers = [
            "ID", "Title", "Severity", "Rule", "Category",
            "Description", "Context", "AI Explanation", 
            "Recommendations", "Status"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=row, column=col)
            cell.value = header
            cell.style = self.styles['header']
        
        row += 1
        
        # Add findings data
        for finding in findings:
            severity = finding.get('severity', 'Medium').upper()
            
            # Prepare cell values
            cell_values = [
                finding.get('id', ''),
                finding.get('title', ''),
                severity,
                finding.get('rule_name', ''),
                finding.get('category', ''),
                finding.get('description', ''),
                json.dumps(finding.get('context', {}), indent=2, default=str) if finding.get('context') else '',
                finding.get('ai_explanation', '') if config.include_ai_explanations else '',
                '\n'.join(finding.get('recommendations', [])) if finding.get('recommendations') else '',
                finding.get('status', 'Open')
            ]
            
            # Add row data
            for col, value in enumerate(cell_values, 1):
                cell = sheet.cell(row=row, column=col, value=value)
                cell.style = self.styles['normal']
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                # Special handling for certain columns
                if col == 3:  # Severity column
                    cell.style = self._get_severity_style(severity)
                elif col in [7, 8]:  # Context and AI Explanation columns
                    cell.style = self.styles['code']
                elif col == 9:  # Recommendations column
                    if value:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            row += 1
            
            # Check row limit
            if row > config.excel_max_rows_per_sheet:
                logger.warning(f"Exceeded maximum rows per sheet ({config.excel_max_rows_per_sheet})")
                sheet[f'A{row}'] = f"... {len(findings) - (row - 4)} more findings not shown"
                break
        
        # Auto-adjust column widths
        for i, width in enumerate([10, 40, 12, 20, 15, 40, 30, 40, 40, 12], 1):
            sheet.column_dimensions[get_column_letter(i)].width = width
        
        # Freeze header row
        sheet.freeze_panes = 'A4'
    
    def _create_recommendations_sheet(self, report_data: Dict[str, Any]):
        """Create recommendations sheet."""
        sheet = self.workbook.create_sheet(title="Recommendations")
        
        # Title
        sheet['A1'] = "Recommendations"
        sheet['A1'].style = self.styles['title']
        sheet.merge_cells('A1:E1')
        
        row = 3
        
        recommendations = report_data.get('recommendations', [])
        
        if not recommendations:
            sheet['A3'] = "No recommendations to report."
            sheet['A3'].style = self.styles['normal']
            return
        
        # Table headers
        headers = ["#", "Title", "Priority", "Timeline", "Actions", "Owner"]
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=row, column=col)
            cell.value = header
            cell.style = self.styles['header']
        
        row += 1
        
        # Add recommendations data
        for i, rec in enumerate(recommendations, 1):
            priority = rec.get('priority', 'Medium').upper()
            
            # Prepare cell values
            cell_values = [
                i,
                rec.get('title', f'Recommendation {i}'),
                priority,
                rec.get('timeline', '30 days'),
                '\n'.join(rec.get('actions', [])),
                rec.get('owner', '')
            ]
            
            # Add row data
            for col, value in enumerate(cell_values, 1):
                cell = sheet.cell(row=row, column=col, value=value)
                cell.style = self.styles['normal']
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                # Apply priority style
                if col == 3:  # Priority column
                    cell.style = self._get_priority_style(priority)
            
            row += 1
        
        # Auto-adjust column widths
        for i, width in enumerate([5, 40, 12, 15, 50, 20], 1):
            sheet.column_dimensions[get_column_letter(i)].width = width
        
        # Freeze header row
        sheet.freeze_panes = 'A4'
    
    def _create_appendix_sheet(self, report_data: Dict[str, Any]):
        """Create appendix sheet."""
        sheet = self.workbook.create_sheet(title="Appendix")
        
        # Title
        sheet['A1'] = "Appendix"
        sheet['A1'].style = self.styles['title']
        sheet.merge_cells('A1:B1')
        
        row = 3
        
        appendix = report_data.get('appendix', {})
        
        if not appendix:
            sheet['A3'] = "No appendix data."
            sheet['A3'].style = self.styles['normal']
            return
        
        for section_title, section_content in appendix.items():
            # Section title
            sheet[f'A{row}'] = section_title
            sheet[f'A{row}'].font = Font(bold=True, size=11, color='2C3E50')
            sheet.merge_cells(f'A{row}:B{row}')
            row += 1
            
            # Section content
            if isinstance(section_content, str):
                sheet[f'A{row}'] = section_content
                sheet[f'A{row}'].style = self.styles['normal']
                sheet.merge_cells(f'A{row}:B{row}')
                row += 1
            elif isinstance(section_content, list):
                for item in section_content:
                    sheet[f'A{row}'] = f"• {item}"
                    sheet[f'A{row}'].style = self.styles['normal']
                    sheet.merge_cells(f'A{row}:B{row}')
                    row += 1
            elif isinstance(section_content, dict):
                for key, value in section_content.items():
                    sheet[f'A{row}'] = f"{key}:"
                    sheet[f'A{row}'].font = Font(bold=True)
                    sheet[f'B{row}'] = str(value)
                    sheet[f'B{row}'].style = self.styles['normal']
                    row += 1
            
            row += 1  # Add spacing between sections
    
    def _create_dashboard_sheet(self, report_data: Dict[str, Any]):
        """Create summary dashboard with charts."""
        sheet = self.workbook.create_sheet(title="Dashboard")
        
        # Title
        sheet['A1'] = "Audit Dashboard"
        sheet['A1'].style = self.styles['title']
        sheet.merge_cells('A1:H1')
        
        row = 3
        
        # Summary metrics
        metrics = report_data.get('metrics', {})
        
        if metrics:
            # Create metric boxes
            metric_data = [
                ("Total Findings", metrics.get('total_findings', 0), "A3"),
                ("Critical", metrics.get('critical', 0), "C3"),
                ("High", metrics.get('high', 0), "E3"),
                ("Medium", metrics.get('medium', 0), "G3"),
                ("Low", metrics.get('low', 0), "I3")
            ]
            
            for label, value, cell_ref in metric_data:
                cell = sheet[cell_ref]
                cell.value = label
                cell.font = Font(bold=True, size=10)
                cell.alignment = Alignment(horizontal='center')
                
                value_cell = sheet[f'{cell_ref[0]}{int(cell_ref[1:]) + 1}']
                value_cell.value = value
                value_cell.font = Font(size=14, bold=True)
                value_cell.alignment = Alignment(horizontal='center')
                
                # Apply severity color for finding counts
                if label in ['Critical', 'High', 'Medium', 'Low']:
                    value_cell.style = self._get_severity_style(label.upper())
                else:
                    value_cell.font = Font(size=14, bold=True, color='2C3E50')
            
            row = 6
        
        # Create severity distribution chart
        findings = report_data.get('findings', [])
        if findings:
            # Count by severity
            severity_counts = {'CRITICAL': 0, 'HIGH': 0, 'MEDIUM': 0, 'LOW': 0}
            for finding in findings:
                severity = finding.get('severity', 'MEDIUM').upper()
                if severity in severity_counts:
                    severity_counts[severity] += 1
            
            # Add data for chart
            sheet[f'A{row}'] = "Severity"
            sheet[f'B{row}'] = "Count"
            sheet[f'A{row}'].style = self.styles['header']
            sheet[f'B{row}'].style = self.styles['header']
            
            row += 1
            
            for severity, count in severity_counts.items():
                if count > 0:
                    sheet[f'A{row}'] = severity.title()
                    sheet[f'B{row}'] = count
                    row += 1
            
            # Create pie chart
            chart = PieChart()
            labels = Reference(sheet, min_col=1, min_row=row - len(severity_counts), max_row=row - 1)
            data = Reference(sheet, min_col=2, min_row=row - len(severity_counts) - 1, max_row=row - 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            chart.title = "Findings by Severity"
            
            # Position chart
            sheet.add_chart(chart, "D3")
            
            # Create bar chart for top rules
            if 'rule_name' in findings[0]:
                # Count by rule
                rule_counts = {}
                for finding in findings[:20]:  # Top 20 rules
                    rule_name = finding.get('rule_name', 'Unknown')
                    rule_counts[rule_name] = rule_counts.get(rule_name, 0) + 1
                
                # Sort by count
                sorted_rules = sorted(rule_counts.items(), key=lambda x: x[1], reverse=True)[:10]
                
                chart_row = row + 2
                sheet[f'A{chart_row}'] = "Rule"
                sheet[f'B{chart_row}'] = "Findings"
                sheet[f'A{chart_row}'].style = self.styles['header']
                sheet[f'B{chart_row}'].style = self.styles['header']
                
                chart_row += 1
                
                for rule_name, count in sorted_rules:
                    sheet[f'A{chart_row}'] = rule_name[:30]  # Truncate long names
                    sheet[f'B{chart_row}'] = count
                    chart_row += 1
                
                # Create bar chart
                bar_chart = BarChart()
                bar_chart.type = "col"
                bar_chart.style = 10
                bar_chart.title = "Top Rules by Finding Count"
                bar_chart.y_axis.title = "Number of Findings"
                bar_chart.x_axis.title = "Rule"
                
                labels = Reference(sheet, min_col=1, min_row=chart_row - len(sorted_rules), max_row=chart_row - 1)
                data = Reference(sheet, min_col=2, min_row=chart_row - len(sorted_rules) - 1, max_row=chart_row - 1)
                bar_chart.add_data(data, titles_from_data=True)
                bar_chart.set_categories(labels)
                
                # Position chart
                sheet.add_chart(bar_chart, "D15")
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            adjusted_width = min(max_length + 2, 30)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _get_severity_style(self, severity: str) -> NamedStyle:
        """Get style for severity level."""
        severity = severity.upper()
        
        if severity == 'CRITICAL':
            return self.styles['critical']
        elif severity == 'HIGH':
            return self.styles['high']
        elif severity == 'MEDIUM':
            return self.styles['medium']
        elif severity == 'LOW':
            return self.styles['low']
        else:
            return self.styles['normal']
    
    def _get_priority_style(self, priority: str) -> NamedStyle:
        """Get style for priority level."""
        priority = priority.upper()
        
        if priority == 'HIGH':
            return self.styles['critical']  # Reuse critical style
        elif priority == 'MEDIUM':
            return self.styles['medium']
        elif priority == 'LOW':
            return self.styles['low']
        else:
            return self.styles['normal']


# Global Excel generator instance
excel_generator = ExcelGenerator()